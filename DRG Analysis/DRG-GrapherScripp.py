import numpy as np
import pandas as pd
import os
#from scipy import stats as scistat
import scipy.optimize as opt
from scipy.stats import sem
from scipy.stats import ttest_ind
from plotnine import *
from openpyxl import load_workbook
from functools import reduce
import csv

"""#Cleaning DRG groups
path = r'D:\Dropbox (Personal)\FileTransfers\DRG'
t1paths = ['9-15_T1', '9-21_T1', '9-23_T1', '9-29_T1']
t2paths = ['9-15_T2', '9-21_T2', '9-23_T2', '9-29_T2']
xlfiles = ['C1-HAC-TC.xlsx', 'Events.xlsx']
atfiles = ['C2Tau.atf','C3Tau.atf']

print('test')
#for n, i in enumerate(t1paths):
#xls = load_workbook(path + '\\' + i + '\\' + xlfiles[0])
fpath = path + '\\' + t1paths[2] + '\\' + xlfiles[0]
xls = load_workbook(fpath)
sheit = xls.sheetnames[0]
print(xls.sheetnames)

#C1-HAC-TC
df = pd.read_excel(fpath, sheit)
df = df[~df['File Name'].str.contains('T2')]
writer = pd.ExcelWriter(fpath, engine='openpyxl')
df.to_excel(writer, sheit)
writer.save()
writer.close()

#Events
fpath = path + '\\' + t1paths[1] + '\\' + xlfiles[1]
xls = load_workbook(fpath)
sheits = xls.sheetnames
for i in sheits:
    if '_T2' in i:
        print(i)
        del xls[i]
writer = pd.ExcelWriter(fpath, engine='openpyxl')
writer.book = xls
writer.save()
writer.close()
#Events xlsx to CSV:
cols = ['Trace', 'Search', 'Category', 'State', 'Event Start Time (ms)', 'Event End Time (ms)', 'Baseline (mV)', 'Peak Amp (mV)', 'Time to Peak (ms)', 'Antipeak Amp (mV)', 'Time to Antipeak (ms)', 'Time of Antipeak (ms)', 'Half-width (ms)', 'Half-amplitude (mV)', 'Time to Rise Half-amplitude (ms)', 'Time to Decay Half-amplitude (ms)', 'Rise Tau (ms)', 'Decay Tau (ms)', 'Max Rise Slope (mV/ms)', 'Time to Max Rise Slope (ms)', 'Max Decay Slope (mV/ms)', 'Time to Max Decay Slope (ms)', 'Rise Slope 10% to 90% (mV/ms)', 'Rise Time 10% to 90% (ms)', 'Decay Slope 90% to 10% (mV/ms)', 'Decay Time 90% to 10% (ms)', 'Area (mV · ms)', 'Inst. Freq. (Hz)', 'Interevent Interval (ms)', 'S.D. of Fit', 'Template Match', ' ']
for i in t1paths:
    fpath = path + '\\' + i + '\\' + xlfiles[1]
    opath = path + '\\' + i + '\\'
    xls = load_workbook(fpath)
    sheits = xls.sheetnames
    for ii in sheits:
        xls = pd.ExcelFile(fpath, engine='openpyxl')
        df = pd.read_excel(xls, ii, header=None, names=cols)
        with open(opath + ii + '.atf', 'w', newline='') as csvfile:
            df.to_csv(csvfile, sep='	')
"""

pd.set_option("display.max_columns", None)
path = r'D:\Dropbox (Personal)\FileTransfers\DRG'
compath = r'C:\Users\ckowalski\Dropbox\FileTransfers\DRG\combinedata.xlsx'
xlspath = path + '\\' + 'finaldata.xlsx'
scrippspath = r'D:\Dropbox (Personal)\FileTransfers\DRG' + '\\finaldata.xlsx' #combinedata.xslx???
#xls = pd.ExcelFile(compath, engine='openpyxl')
xls = pd.ExcelFile(scrippspath, engine='openpyxl')
dataramp = pd.read_excel(xls, 'Ramps')
datac1 = pd.read_excel(xls, 'C1')
datac2 = pd.read_excel(xls, 'C2')
datac3 = pd.read_excel(xls, 'C3')

dataramp['CellID'] = dataramp['Date'] + '_' + dataramp.astype(str)['Cell'] + '_' + dataramp['Protocol']
#datarheo = pd.read_excel(xls, 'RheoGB') # Moved below after creating datarheo2
#datafit = pd.read_excel(xls, 'HillFit')
#datac2 = datac23[datac23['Protocol'].str.contains('C2')].copy()
#datac3 = datac23[datac23['Protocol'].str.contains('C3')].copy()

#Switches
bool_fitting        = 0
bool_id_EC50        = 0
bool_id_TAC         = 0
bool_id_HAC         = 0
bool_id_HACBase     = 0
bool_id_C2DecayTau  = 0
bool_id_C3DecayTau  = 0
bool_id_ramptime    = 0
bool_id_ramp        = 1
bool_id_rampfree    = 0
bool_id_Rheo        = 0
bool_id_genRheoBase = 0

#Rheobase:
"""# Groupby ID combineddata + 9-15 -> Min(Stim (pA)) -> subset -> xlsx
path = r'D:\Dropbox (Personal)\FileTransfers\DRG'
scrippspath = r'D:\Dropbox (Personal)\FileTransfers\DRG' + '\\finaldata.xlsx' #combinedata.xslx???
xls = pd.ExcelFile(scrippspath, engine='openpyxl')
dataramp = pd.read_excel(xls, 'Ramps')
datac1 = pd.read_excel(xls, 'C1')
datac2 = pd.read_excel(xls, 'C2')
datac3 = pd.read_excel(xls, 'C3')"""


#Rheobas Contd.
rheodf = dataramp[['Date', 'Genotype', 'Cell', 'CellID', 'Protocol', 'CellType', 'Include', 'Inst. Freq. (Hz)', 'Stim (pA)', 'Event Start Time (ms)']]
rheodf2 = rheodf.groupby('CellID').min().rename(columns={'Stim (pA)':'Rheobase'})


if bool_id_genRheoBase:
    book = load_workbook(scrippspath)
    writer = pd.ExcelWriter(scrippspath, engine='openpyxl')
    writer.book = book
    rheodf.to_excel(writer, 'RheoAll')
    rheodf2.to_excel(writer, 'RheoGB')
    writer.save()
    writer.close()

xls = pd.ExcelFile(scrippspath, engine='openpyxl')
dataramp = pd.read_excel(xls, 'Ramps')
datac1 = pd.read_excel(xls, 'C1')
datac2 = pd.read_excel(xls, 'C2')
datac3 = pd.read_excel(xls, 'C3')
datarheo = pd.read_excel(xls, 'RheoGB')

#Filtering
#Compensate HEKA CC gain:
#dataramp['Stim (pA)'] = dataramp['Stim (pA)']/5
#Drop by Include column:
#dataramporig = dataramp.copy()
dataramp = dataramp[dataramp['Include'] > 0]
#Add CellID for funky free-axis facets
#todo: maybe just make iterator to plot individually...
dataramp['CellID'] = dataramp['Date'] + '_' + dataramp.astype(str)['Cell'] + '_' + dataramp['Protocol']
#dataramp = dataramp.replace('RampTreat', 'RampBase').replace('RampBaseline', 'RampTreat')
#Convert dates
dataramp['Date'] = dataramp['Date'].str.replace('9-15_T1', '9-15').str.replace('9-15_T2', '9-15').str.replace('9-21_T1', '9-21').str.replace('9-21_T2', '9-21').str.replace('9-23_T1', '9-23').str.replace('9-23_T2', '9-23').str.replace('9-29_T1', '9-29').str.replace('9-29_T2', '9-29')
datac1['Date'] = datac1['Date'].str.replace('9-15_T1', '9-15').str.replace('9-15_T2', '9-15').str.replace('9-21_T1', '9-21').str.replace('9-21_T2', '9-21').str.replace('9-23_T1', '9-23').str.replace('9-23_T2', '9-23').str.replace('9-29_T1', '9-29').str.replace('9-29_T2', '9-29')
datac2['Date'] = datac2['Date'].str.replace('9-15_T1', '9-15').str.replace('9-15_T2', '9-15').str.replace('9-21_T1', '9-21').str.replace('9-21_T2', '9-21').str.replace('9-23_T1', '9-23').str.replace('9-23_T2', '9-23').str.replace('9-29_T1', '9-29').str.replace('9-29_T2', '9-29')
datac3['Date'] = datac3['Date'].str.replace('9-15_T1', '9-15').str.replace('9-15_T2', '9-15').str.replace('9-21_T1', '9-21').str.replace('9-21_T2', '9-21').str.replace('9-23_T1', '9-23').str.replace('9-23_T2', '9-23').str.replace('9-29_T1', '9-29').str.replace('9-29_T2', '9-29')
# Remove dates for subplotting
dataramp = dataramp[(dataramp['Date']=='9-21') | (dataramp['Date']=='9-15')]
datac1 = datac1[(datac1['Date']=='9-15') | (datac1['Date']=='9-21')]
datac2 = datac2[(datac2['Date']=='9-15') | (datac2['Date']=='9-21')]
datac3 = datac3[(datac3['Date']=='9-15') | (datac3['Date']=='9-21')]


#Rheobase Stats:
# Groupby ID combineddata + 9-15 -> Min(Stim (pA)) -> subset -> xlsx
#datarheo = datarheo.drop(datarheo[datarheo['Protocol']=='RampRetreat'].index)
#datarheo = datarheo.drop(datarheo[datarheo['Protocol']=='RampRerecov'].index)
datarheo['Group'] = datarheo.apply(lambda x: '%s_%s_%s' % (x['Genotype'], x['Cell'], x['Protocol']), axis=1) #String concat
datarheoMean = datarheo.groupby('Group').mean()
#datarheoSEM = datarheo.groupby('Group').apply(lambda x: sem(x))
datarheoSEM = datarheo.groupby('Group').sem()

datarheo['Protocol'] = pd.Categorical(datarheo.Protocol, categories=['RampBaseline', 'RampTreat', 'RampRecov', 'RampRetreat', 'RampRerecov'])
datarheo['Group'] = pd.Categorical(datarheo.Group, categories=['CtrlRampBaseline', 'CtrlRampTreat', 'CtrlRampRecov', 'KORampBaseline', 'KORampTreat', 'KORampRecov'])

datarheostat = pd.DataFrame.merge(datarheoMean, datarheoSEM, on='Group', how='outer', suffixes=('_mean', '_sem')).reset_index()
datarheostat['Group'] = pd.Categorical(datarheostat.Group, categories=['CtrlRampBaseline', 'CtrlRampTreat', 'CtrlRampRecov', 'KORampBaseline', 'KORampTreat', 'KORampRecov'])

print('Ctrl Baseline vs. Treat:', '\n', ttest_ind(datarheo[datarheo['Group']=='CtrlRampBaseline']['Rheobase'], datarheo[datarheo['Group']=='CtrlRampTreat']['Rheobase']))
print('KO Baseline vs. Treat:', '\n', ttest_ind(datarheo[datarheo['Group']=='KORampBaseline']['Rheobase'], datarheo[datarheo['Group']=='KORampTreat']['Rheobase']))
print('Ctrl Treat vs. KO Treat:', '\n', ttest_ind(datarheo[datarheo['Group']=='CtrlRampTreat']['Rheobase'], datarheo[datarheo['Group']=='KORampTreat']['Rheobase']))

#Curve Fitting
def ll4(x,b,c,d,e):
    '''This function is basically a copy of the LL.4 function from the R drc package with
     - b: hill slope
     - c: min response
     - d: max response
     - e: EC50'''
    return(c+(d-c)/(1+np.exp(b*(np.log(x)-np.log(e)))))

if bool_fitting:
    RampsList = []
    for i in dataramp['ID'].unique():
        RampsList.append(i)
    datafit = pd.DataFrame()
    for i in RampsList:
        print(i)
        rampslice = dataramp[dataramp['ID']==i]
        initc = rampslice['Inst. Freq. (Hz)'].min()
        initd = rampslice['Inst. Freq. (Hz)'].max()
        inite = rampslice['Stim (pA)'].mean()
        initials = [0.05, initc, initd, inite]
        fitco, covm = opt.curve_fit(ll4, rampslice['Stim (pA)'], rampslice['Inst. Freq. (Hz)'], p0=initials, maxfev=10000)
        rampslice['RawResiduals'] = rampslice['Inst. Freq. (Hz)'].apply(lambda x: ll4(x,*fitco))
        curfit = dict(zip(['b', 'c', 'd', 'e'], fitco))
        curfit['ID']=i
        curfit['Residual'] = sum(rampslice['RawResiduals']**2)
        for key in curfit.keys():
            rampslice[key] = curfit[key]
        datafit = datafit.append(rampslice)

    book = load_workbook(path + '\\finaldata.xlsx')
    writer = pd.ExcelWriter(path + '\\finaldata.xlsx', engine='openpyxl')
    writer.book = book
    datafit.to_excel(writer, 'HillFit')
    writer.save()
    writer.close()


if bool_id_EC50:
    id_EC50 = (ggplot(data=datafit, mapping=aes(x='Cell', y='e', color='Protocol'))
                    + geom_point(size=1)
                    + facet_grid('Date ~ .', space='free')
                    #+ theme_light()
                    + theme(aspect_ratio=1)
                    + theme(subplots_adjust={'right': 0.75})
                    + theme(strip_text_y= element_text(angle = 0, ha = 'left')))
    fig = id_EC50.draw()
    #fig.set_size_inches(9, 108, forward=True)
    #group_stimfreq.draw(show=True)
    fig.savefig(str(path+ '\\id_ec50.png'), dpi=300)

if bool_id_TAC:
    id_TAC = (ggplot(data=datac1, mapping=aes(x='Trace', y='TACAntipeak', color='Protocol'))
                    + geom_line(size=1)
                    + facet_grid('Cell ~ Date', space='free')
                    #+ theme_light()
                    + theme(aspect_ratio=1)
                    + theme(subplots_adjust={'right': 0.75})
                    + theme(strip_text_y= element_text(angle = 0, ha = 'left')))
    fig = id_TAC.draw()
    #fig.set_size_inches(9, 108, forward=True)
    #group_stimfreq.draw(show=True)
    fig.savefig(str(path+ '\\id_TAC.png'), dpi=300)

if bool_id_HAC:
    id_HAC = (ggplot(data=datac1, mapping=aes(x='Trace', y='HACDiff', color='Protocol'))
                    + geom_line(size=1)
                    + facet_grid('Cell ~ Date', space='free')
                    #+ theme_light()
                    + theme(aspect_ratio=1)
                    + theme(subplots_adjust={'right': 0.75})
                    + theme(strip_text_y= element_text(angle = 0, ha = 'left')))
    fig = id_HAC.draw()
    #fig.set_size_inches(9, 108, forward=True)
    #group_stimfreq.draw(show=True)
    fig.savefig(str(path+ '\\id_HAC.png'), dpi=300)

if bool_id_HACBase:
    id_HACbase = (ggplot(data=datac1, mapping=aes(x='Trace', y='HACBase', color='Protocol'))
              + geom_line(size=1)
              + facet_grid('Cell ~ Date', space='free')
              # + theme_light()
              + theme(aspect_ratio=1)
              + theme(subplots_adjust={'right': 0.75})
              + theme(strip_text_y=element_text(angle=0, ha='left')))
    fig = id_HACbase.draw()
    # fig.set_size_inches(9, 108, forward=True)
    # group_stimfreq.draw(show=True)
    fig.savefig(str(path + '\\id_HACbase.png'), dpi=300)

if bool_id_C2DecayTau:
    id_C2DecayTau = (ggplot(data=datac2, mapping=aes(x='A', y='tau', color='Protocol'))
                    + geom_point(size=1)
                    + facet_grid('Date ~ Cell', space='free')
                    #+ theme_light()
                    + theme(aspect_ratio=1)
                    + theme(subplots_adjust={'right': 0.75})
                    + theme(strip_text_y= element_text(angle = 0, ha = 'left')))
    fig = id_C2DecayTau.draw()
    #fig.set_size_inches(9, 108, forward=True)
    #group_stimfreq.draw(show=True)
    fig.savefig(str(path+ '\\id_C2DecayTauByIntercept.png'), dpi=300)

if bool_id_C3DecayTau:
    id_C3DecayTau = (ggplot(data=datac3, mapping=aes(x='tau1', y='tau2', color='Protocol'))
                    + geom_point(size=1)
                    + facet_grid('Date ~ Cell', space='free')
                    #+ theme_light()
                    + theme(aspect_ratio=1)
                    + theme(subplots_adjust={'right': 0.75})
                    + theme(strip_text_y= element_text(angle = 0, ha = 'left')))
    fig = id_C3DecayTau.draw()
    #fig.set_size_inches(9, 108, forward=True)
    #group_stimfreq.draw(show=True)
    fig.savefig(str(path+ '\\id_C3DecayTauByTau.png'), dpi=300)

if bool_id_ramp:
    id_ramp = (ggplot(data=dataramp, mapping=aes(x='Stim (pA)', y='Inst. Freq. (Hz)', color='Protocol', xmax=2000))
                    + geom_point(size=.01)
                    + facet_grid('Cell ~ Date', space='free')
                    #+ theme_light()
                    #+ theme(aspect_ratio=2)
                    + theme(subplots_adjust={'right': 0.75})
                    + theme(strip_text_y= element_text(angle = 0, ha = 'left')))
    fig = id_ramp.draw()
    #fig.set_size_inches(9, 108, forward=True)
    #group_stimfreq.draw(show=True)
    fig.savefig(str(path+ '\\id_ramp.png'), dpi=300)

if bool_id_rampfree:
    id_rampfree = (ggplot(data=dataramp, mapping=aes(x='Stim (pA)', y='Inst. Freq. (Hz)', color='Protocol'))
                    + geom_point(size=.05)
                    + facet_wrap('CellID', scales='free')
                    #+ theme_light()
                    #+ theme(aspect_ratio=2)
                    + theme(subplots_adjust={'right': 0.75})
                    + theme(subplots_adjust={'wspace': 0.25})
                    + theme(strip_text_y= element_text(angle = 0, ha = 'left')))
    fig = id_rampfree.draw()
    #fig.set_size_inches(9, 108, forward=True)
    #group_stimfreq.draw(show=True)
    fig.savefig(str(path+ '\\id_rampfree.png'), dpi=300)

if bool_id_ramptime:
    id_ramptime = (ggplot(data=dataramp, mapping=aes(x='Event Start Time (ms)', y='Inst. Freq. (Hz)', color='Protocol'))
                    + geom_point(size=.05)
                    + facet_grid('Cell ~ Date', space='free')
                    #+ theme_light()
                    #+ theme(aspect_ratio=2)
                    + theme(subplots_adjust={'right': 0.75})
                    + theme(strip_text_y= element_text(angle = 0, ha = 'left')))
    fig = id_ramptime.draw()
    #fig.set_size_inches(9, 108, forward=True)
    #group_stimfreq.draw(show=True)
    fig.savefig(str(path+ '\\id_ramptime.png'), dpi=300)

if bool_id_Rheo:
    id_Rheo = (ggplot(data=datarheo, mapping=aes(x='Group', y='Rheobase', color='CellID', ymin=0, ymax=1200))
                    + geom_point(size=1)
#                    + stat_summary(geom = "bar", fun_y = np.mean)
#                    + stat_summary(geom = "errorbar", fun_data = 'mean_se')
                    #+ geom_boxplot()
#                    +geom_bar(aes(x='Group', y='Rheobase'), data=datarheoMean)
#                    +geom_errorbar(aes(x='Group', y='Rheobase'), data=datarheoSEM)
#                    + facet_grid('. ~ Genotype', space='free')
                    #+ theme_light()
#                    + ylim(0, 1200)
                    + theme(aspect_ratio=2/3)
                    + theme(subplots_adjust={'right': .98, 'bottom': 0.25})
                    + theme(strip_text_y= element_text(angle = 0, ha = 'left'))
                    + theme(axis_text_x= element_text(rotation=45, hjust=1)))
    fig = id_Rheo.draw()
    #fig.set_size_inches(9, 108, forward=True)
    #group_stimfreq.draw(show=True)
    fig.savefig(str(path+ '\\id_Rheo.png'), dpi=300)

    mean_Rheo = (ggplot(data=datarheostat, mapping=aes(x='Group', y='Rheobase_mean', color='Group', ymin=0, ymax=1200))
               + geom_point(size=1)
               #                    + stat_summary(geom = "bar", fun_y = np.mean)
               #                    + stat_summary(geom = "errorbar", fun_data = 'mean_se')
               # + geom_boxplot()
                +geom_bar(aes(x='Group', y='Rheobase_mean'), stat='identity')
                +geom_errorbar(aes(x='Group', ymax='Rheobase_mean + Rheobase_sem', ymin='Rheobase_mean - Rheobase_sem'))
                +geom_point(aes(x='Group', y='Rheobase', color='CellID'), data=datarheo)
               #                    +geom_errorbar(aes(x='Group', y='Rheobase'), data=datarheoSEM)
               #                    + facet_grid('. ~ Genotype', space='free')
               # + theme_light()
#               + ylim(0, 1200)
               + theme(aspect_ratio=2 / 3)
               + theme(subplots_adjust={'right': .98, 'bottom': 0.25})
               + theme(strip_text_y=element_text(angle=0, ha='left'))
               + theme(axis_text_x=element_text(rotation=45, hjust=1)))
    fig = mean_Rheo.draw()
    # fig.set_size_inches(9, 108, forward=True)
    # group_stimfreq.draw(show=True)
    fig.savefig(str(path + '\\id_Rheo_Means.png'), dpi=300)