import numpy as np
import pandas as pd
import os
#from scipy import stats as scistat
import scipy.optimize as opt
from scipy.stats import sem
from scipy.stats import ttest_ind
from plotnine import *
from openpyxl import load_workbook
from functools import reduce

pd.set_option("display.max_columns", None)
path = r'D:\Dropbox (Personal)\FileTransfers\DRG'
compath = r'C:\Users\ckowalski\Dropbox\FileTransfers\DRG\combinedata.xlsx'
xlspath = path + '\\' + 'finaldata.xlsx'
scrippspath = r'D:\Dropbox (Personal)\FileTransfers\DRG' + '\\finaldata.xlsx' #combinedata.xslx???
#xls = pd.ExcelFile(compath, engine='openpyxl')
xls = pd.ExcelFile(scrippspath, engine='openpyxl')
dataramp = pd.read_excel(xls, 'Ramps')
datac1 = pd.read_excel(xls, 'C1')
datac2 = pd.read_excel(xls, 'C2')
datac3 = pd.read_excel(xls, 'C3')

dataramp['CellID'] = dataramp['Date'] + '_' + dataramp.astype(str)['Cell'] + '_' + dataramp['Protocol']
#datarheo = pd.read_excel(xls, 'RheoGB') # Moved below after creating datarheo2
#datafit = pd.read_excel(xls, 'HillFit')
#datac2 = datac23[datac23['Protocol'].str.contains('C2')].copy()
#datac3 = datac23[datac23['Protocol'].str.contains('C3')].copy()

#Switches
bool_fitting        = 0
bool_id_EC50        = 0
bool_id_TAC         = 0
bool_id_HAC         = 0
bool_id_HACBase     = 0
bool_id_C2DecayTau  = 0
bool_id_C3DecayTau  = 0
bool_id_ramptime    = 0
bool_id_ramp        = 0
bool_id_rampfree    = 0
bool_id_Rheo        = 0
bool_id_RheoDiff    = 1
bool_id_RheoInhib   = 1
bool_id_genRheoBase = 0

#Rheobase:
"""# Groupby ID combineddata + 9-15 -> Min(Stim (pA)) -> subset -> xlsx
path = r'D:\Dropbox (Personal)\FileTransfers\DRG'
scrippspath = r'D:\Dropbox (Personal)\FileTransfers\DRG' + '\\finaldata.xlsx' #combinedata.xslx???
xls = pd.ExcelFile(scrippspath, engine='openpyxl')
dataramp = pd.read_excel(xls, 'Ramps')
datac1 = pd.read_excel(xls, 'C1')
datac2 = pd.read_excel(xls, 'C2')
datac3 = pd.read_excel(xls, 'C3')"""


#Rheobas Contd.
rheodf = dataramp[['Date', 'Genotype', 'Cell', 'CellID', 'Protocol', 'CellType', 'Include', 'Inst. Freq. (Hz)', 'Stim (pA)', 'Event Start Time (ms)']]
rheodf2 = rheodf.groupby('CellID').min().rename(columns={'Stim (pA)':'Rheobase'})


if bool_id_genRheoBase:
    book = load_workbook(scrippspath)
    writer = pd.ExcelWriter(scrippspath, engine='openpyxl')
    writer.book = book
    rheodf.to_excel(writer, 'RheoAll')
    rheodf2.to_excel(writer, 'RheoGB')
    writer.save()
    writer.close()

xls = pd.ExcelFile(scrippspath, engine='openpyxl')
dataramp = pd.read_excel(xls, 'Ramps')
datac1 = pd.read_excel(xls, 'C1')
datac2 = pd.read_excel(xls, 'C2')
datac3 = pd.read_excel(xls, 'C3')
datarheo = pd.read_excel(xls, 'RheoGB')

#Filtering
#Compensate HEKA CC gain:
dataramp['Stim (pA)'] = dataramp['Stim (pA)']/5
#Drop by Include column:
#dataramporig = dataramp.copy()
dataramp = dataramp[dataramp['Include'] > 0]
#Add CellID for funky free-axis facets
#todo: maybe just make iterator to plot individually...
dataramp['CellID'] = dataramp['Date'] + '_' + dataramp.astype(str)['Cell'] + '_' + dataramp['Protocol']
#dataramp = dataramp.replace('RampTreat', 'RampBase').replace('RampBaseline', 'RampTreat')
#Convert dates
dataramp['Date'] = dataramp['Date'].str.replace('9-15_T1', '9-15').str.replace('9-15_T2', '9-15').str.replace('9-21_T1', '9-21').str.replace('9-21_T2', '9-21').str.replace('9-23_T1', '9-23').str.replace('9-23_T2', '9-23').str.replace('9-29_T1', '9-29').str.replace('9-29_T2', '9-29')
datac1['Date'] = datac1['Date'].str.replace('9-15_T1', '9-15').str.replace('9-15_T2', '9-15').str.replace('9-21_T1', '9-21').str.replace('9-21_T2', '9-21').str.replace('9-23_T1', '9-23').str.replace('9-23_T2', '9-23').str.replace('9-29_T1', '9-29').str.replace('9-29_T2', '9-29')
datac2['Date'] = datac2['Date'].str.replace('9-15_T1', '9-15').str.replace('9-15_T2', '9-15').str.replace('9-21_T1', '9-21').str.replace('9-21_T2', '9-21').str.replace('9-23_T1', '9-23').str.replace('9-23_T2', '9-23').str.replace('9-29_T1', '9-29').str.replace('9-29_T2', '9-29')
datac3['Date'] = datac3['Date'].str.replace('9-15_T1', '9-15').str.replace('9-15_T2', '9-15').str.replace('9-21_T1', '9-21').str.replace('9-21_T2', '9-21').str.replace('9-23_T1', '9-23').str.replace('9-23_T2', '9-23').str.replace('9-29_T1', '9-29').str.replace('9-29_T2', '9-29')
# Remove dates for subplotting
dataramp = dataramp[(dataramp['Date']=='9-21') | (dataramp['Date']=='9-15')]
datac1 = datac1[(datac1['Date']=='9-15') | (datac1['Date']=='9-21')]
datac2 = datac2[(datac2['Date']=='9-15') | (datac2['Date']=='9-21')]
datac3 = datac3[(datac3['Date']=='9-15') | (datac3['Date']=='9-21')]


#Rheobase Stats:
# Groupby ID combineddata + 9-15 -> Min(Stim (pA)) -> subset -> xlsx
#datarheo = datarheo.drop(datarheo[datarheo['Protocol']=='RampRetreat'].index)
#datarheo = datarheo.drop(datarheo[datarheo['Protocol']=='RampRerecov'].index)
#datarheo = datarheo.drop(datarheo[datarheo['Protocol']=='RampRecov'].index)
datarheo['Group'] = datarheo.apply(lambda x: '%s_%s_%s' % (x['Genotype'], x['Cell'], x['Protocol']), axis=1) #String concat


# %Inhib
datarheoInhib = dict()
datarheoDiff = dict()
#for i in datarheo['CellID.1']:
#    ldf = datarheo[datarheo['CellID.1']==i]
#    ldf[ldf['Protocol']=='RampBaseline']

for i in datarheo['CellID']:
    try:
        datarheoInhib[i] = datarheo[datarheo['CellID']==i][datarheo['Protocol']=='RampBaseline']['Rheobase'].iloc[0]/datarheo[datarheo['CellID'] == i][datarheo['Protocol']=='RampTreat']['Rheobase'].iloc[0]
        datarheoDiff[i] = datarheo[datarheo['CellID']==i][datarheo['Protocol']=='RampTreat']['Rheobase'].iloc[0] - datarheo[datarheo['CellID']==i][datarheo['Protocol']=='RampBaseline']['Rheobase'].iloc[0]
    except Exception as e:
        print('exception:', e, 'cell:', datarheo[datarheo['CellID']==i])

datarheoInhibDF = pd.DataFrame.from_dict(datarheoInhib, orient='index', columns=['RheoInhibition']).reset_index().rename(columns={'index':'CellID'})
datarheoInhibDF['RheoInhibition'] = datarheoInhibDF['RheoInhibition'].apply(lambda x: (1-x)*100)
datarheoDiffDF = pd.DataFrame.from_dict(datarheoDiff, orient='index', columns=['RheoDiff']).reset_index().rename(columns={'index':'CellID'})
datarheodfs = pd.merge(datarheoDiffDF, datarheoInhibDF, how='outer', on='CellID')
datarheo = pd.merge(datarheo, datarheodfs, how='outer', on='CellID')

datarheoMean = datarheo.groupby('Group').mean()
#datarheoSEM = datarheo.groupby('Group').apply(lambda x: sem(x))
datarheoSEM = datarheo.groupby('Group').sem()

datarheo['Protocol'] = pd.Categorical(datarheo.Protocol, categories=['RampBaseline', 'RampTreat', 'RampRecov', 'RampRetreat', 'RampRerecov'])
datarheo['Group'] = pd.Categorical(datarheo.Group, categories=['CtrlRampBaseline', 'CtrlRampTreat', 'CtrlRampRecov', 'KORampBaseline', 'KORampTreat', 'KORampRecov'])

datarheostat = pd.DataFrame.merge(datarheoMean, datarheoSEM, on='Group', how='outer', suffixes=('_mean', '_sem')).reset_index()
datarheostat['Group'] = pd.Categorical(datarheostat.Group, categories=['CtrlRampBaseline', 'CtrlRampTreat', 'CtrlRampRecov', 'KORampBaseline', 'KORampTreat', 'KORampRecov'])

print('Ctrl Baseline vs. Treat:', '\n', ttest_ind(datarheo[datarheo['Group']=='CtrlRampBaseline']['Rheobase'], datarheo[datarheo['Group']=='CtrlRampTreat']['Rheobase']))
print('KO Baseline vs. Treat:', '\n', ttest_ind(datarheo[datarheo['Group']=='KORampBaseline']['Rheobase'], datarheo[datarheo['Group']=='KORampTreat']['Rheobase']))
print('Ctrl Treat vs. KO Treat:', '\n', ttest_ind(datarheo[datarheo['Group']=='CtrlRampTreat']['Rheobase'], datarheo[datarheo['Group']=='KORampTreat']['Rheobase']))
print('Inhib:', '\n', ttest_ind(datarheo[datarheo['Group']=='CtrlRampTreat']['RheoInhibition'], datarheo[datarheo['Group']=='KORampTreat']['RheoInhibition']))
print('Diff:', '\n', ttest_ind(datarheo[datarheo['Group']=='CtrlRampTreat']['RheoDiff'], datarheo[datarheo['Group']=='KORampTreat']['RheoDiff']))
datarheostat = datarheostat.drop(datarheostat[datarheostat['Group']=='CtrlRampBaseline'].index)
datarheostat = datarheostat.drop(datarheostat[datarheostat['Group']=='KORampBaseline'].index)
datarheo = datarheo.drop(datarheo[datarheo['Group']=='CtrlRampBaseline'].index)
datarheo = datarheo.drop(datarheo[datarheo['Group']=='KORampBaseline'].index)



#Curve Fitting
def ll4(x,b,c,d,e):
    '''This function is basically a copy of the LL.4 function from the R drc package with
     - b: hill slope
     - c: min response
     - d: max response
     - e: EC50'''
    return(c+(d-c)/(1+np.exp(b*(np.log(x)-np.log(e)))))

if bool_fitting:
    RampsList = []
    for i in dataramp['ID'].unique():
        RampsList.append(i)
    datafit = pd.DataFrame()
    for i in RampsList:
        print(i)
        rampslice = dataramp[dataramp['ID']==i]
        initc = rampslice['Inst. Freq. (Hz)'].min()
        initd = rampslice['Inst. Freq. (Hz)'].max()
        inite = rampslice['Stim (pA)'].mean()
        initials = [0.05, initc, initd, inite]
        fitco, covm = opt.curve_fit(ll4, rampslice['Stim (pA)'], rampslice['Inst. Freq. (Hz)'], p0=initials, maxfev=10000)
        rampslice['RawResiduals'] = rampslice['Inst. Freq. (Hz)'].apply(lambda x: ll4(x,*fitco))
        curfit = dict(zip(['b', 'c', 'd', 'e'], fitco))
        curfit['ID']=i
        curfit['Residual'] = sum(rampslice['RawResiduals']**2)
        for key in curfit.keys():
            rampslice[key] = curfit[key]
        datafit = datafit.append(rampslice)

    book = load_workbook(path + '\\finaldata.xlsx')
    writer = pd.ExcelWriter(path + '\\finaldata.xlsx', engine='openpyxl')
    writer.book = book
    datafit.to_excel(writer, 'HillFit')
    writer.save()
    writer.close()


if bool_id_EC50:
    id_EC50 = (ggplot(data=datafit, mapping=aes(x='Cell', y='e', color='Protocol'))
                    + geom_point(size=1)
                    + facet_grid('Date ~ .', space='free')
                    #+ theme_light()
                    + theme(aspect_ratio=1)
                    + theme(subplots_adjust={'right': 0.75})
                    + theme(strip_text_y= element_text(angle = 0, ha = 'left')))
    fig = id_EC50.draw()
    #fig.set_size_inches(9, 108, forward=True)
    #group_stimfreq.draw(show=True)
    fig.savefig(str(path+ '\\id_ec50.png'), dpi=300)

if bool_id_TAC:
    id_TAC = (ggplot(data=datac1, mapping=aes(x='Trace', y='TACAntipeak', color='Protocol'))
                    + geom_line(size=1)
                    + facet_grid('Cell ~ Date', space='free')
                    #+ theme_light()
                    + theme(aspect_ratio=1)
                    + theme(subplots_adjust={'right': 0.75})
                    + theme(strip_text_y= element_text(angle = 0, ha = 'left')))
    fig = id_TAC.draw()
    #fig.set_size_inches(9, 108, forward=True)
    #group_stimfreq.draw(show=True)
    fig.savefig(str(path+ '\\id_TAC.png'), dpi=300)

if bool_id_HAC:
    id_HAC = (ggplot(data=datac1, mapping=aes(x='Trace', y='HACDiff', color='Protocol'))
                    + geom_line(size=1)
                    + facet_grid('Cell ~ Date', space='free')
                    #+ theme_light()
                    + theme(aspect_ratio=1)
                    + theme(subplots_adjust={'right': 0.75})
                    + theme(strip_text_y= element_text(angle = 0, ha = 'left')))
    fig = id_HAC.draw()
    #fig.set_size_inches(9, 108, forward=True)
    #group_stimfreq.draw(show=True)
    fig.savefig(str(path+ '\\id_HAC.png'), dpi=300)

if bool_id_HACBase:
    id_HACbase = (ggplot(data=datac1, mapping=aes(x='Trace', y='HACBase', color='Protocol'))
              + geom_line(size=1)
              + facet_grid('Cell ~ Date', space='free')
              # + theme_light()
              + theme(aspect_ratio=1)
              + theme(subplots_adjust={'right': 0.75})
              + theme(strip_text_y=element_text(angle=0, ha='left')))
    fig = id_HACbase.draw()
    # fig.set_size_inches(9, 108, forward=True)
    # group_stimfreq.draw(show=True)
    fig.savefig(str(path + '\\id_HACbase.png'), dpi=300)

if bool_id_C2DecayTau:
    id_C2DecayTau = (ggplot(data=datac2, mapping=aes(x='A', y='tau', color='Protocol'))
                    + geom_point(size=1)
                    + facet_grid('Date ~ Cell', space='free')
                    #+ theme_light()
                    + theme(aspect_ratio=1)
                    + theme(subplots_adjust={'right': 0.75})
                    + theme(strip_text_y= element_text(angle = 0, ha = 'left')))
    fig = id_C2DecayTau.draw()
    #fig.set_size_inches(9, 108, forward=True)
    #group_stimfreq.draw(show=True)
    fig.savefig(str(path+ '\\id_C2DecayTauByIntercept.png'), dpi=300)

if bool_id_C3DecayTau:
    id_C3DecayTau = (ggplot(data=datac3, mapping=aes(x='tau1', y='tau2', color='Protocol'))
                    + geom_point(size=1)
                    + facet_grid('Date ~ Cell', space='free')
                    #+ theme_light()
                    + theme(aspect_ratio=1)
                    + theme(subplots_adjust={'right': 0.75})
                    + theme(strip_text_y= element_text(angle = 0, ha = 'left')))
    fig = id_C3DecayTau.draw()
    #fig.set_size_inches(9, 108, forward=True)
    #group_stimfreq.draw(show=True)
    fig.savefig(str(path+ '\\id_C3DecayTauByTau.png'), dpi=300)

if bool_id_ramp:
    id_ramp = (ggplot(data=dataramp, mapping=aes(x='Stim (pA)', y='Inst. Freq. (Hz)', color='Protocol'))
                    + geom_point(size=.05)
                    + facet_grid('Cell ~ Date', space='free')
                    #+ theme_light()
                    #+ theme(aspect_ratio=2)
                    + theme(subplots_adjust={'right': 0.75})
                    + theme(strip_text_y= element_text(angle = 0, ha = 'left')))
    fig = id_ramp.draw()
    #fig.set_size_inches(9, 108, forward=True)
    #group_stimfreq.draw(show=True)
    fig.savefig(str(path+ '\\id_ramp.png'), dpi=300)

if bool_id_rampfree:
    id_rampfree = (ggplot(data=dataramp, mapping=aes(x='Stim (pA)', y='Inst. Freq. (Hz)', color='Protocol'))
                    + geom_point(size=.05)
                    + facet_wrap('CellID', scales='free')
                    #+ theme_light()
                    #+ theme(aspect_ratio=2)
                    + theme(subplots_adjust={'right': 0.75})
                    + theme(subplots_adjust={'wspace': 0.25})
                    + theme(strip_text_y= element_text(angle = 0, ha = 'left')))
    fig = id_rampfree.draw()
    #fig.set_size_inches(9, 108, forward=True)
    #group_stimfreq.draw(show=True)
    fig.savefig(str(path+ '\\id_rampfree.png'), dpi=300)

if bool_id_ramptime:
    id_ramptime = (ggplot(data=dataramp, mapping=aes(x='Event Start Time (ms)', y='Inst. Freq. (Hz)', color='Protocol'))
                    + geom_point(size=.05)
                    + facet_grid('Cell ~ Date', space='free')
                    #+ theme_light()
                    #+ theme(aspect_ratio=2)
                    + theme(subplots_adjust={'right': 0.75})
                    + theme(strip_text_y= element_text(angle = 0, ha = 'left')))
    fig = id_ramptime.draw()
    #fig.set_size_inches(9, 108, forward=True)
    #group_stimfreq.draw(show=True)
    fig.savefig(str(path+ '\\id_ramptime.png'), dpi=300)

if bool_id_Rheo:
    id_Rheo = (ggplot(data=datarheo, mapping=aes(x='Group', y='Rheobase', color='CellID', ymin=0, ymax=1200))
                    + geom_point(size=1)
#                    + stat_summary(geom = "bar", fun_y = np.mean)
#                    + stat_summary(geom = "errorbar", fun_data = 'mean_se')
                    #+ geom_boxplot()
#                    +geom_bar(aes(x='Group', y='Rheobase'), data=datarheoMean)
#                    +geom_errorbar(aes(x='Group', y='Rheobase'), data=datarheoSEM)
#                    + facet_grid('. ~ Genotype', space='free')
                    #+ theme_light()
#                    + ylim(0, 1200)
                    + theme(aspect_ratio=2/3)
                    + theme(subplots_adjust={'right': .98, 'bottom': 0.25})
                    + theme(strip_text_y= element_text(angle = 0, ha = 'left'))
                    + theme(axis_text_x= element_text(rotation=45, hjust=1)))
    fig = id_Rheo.draw()
    #fig.set_size_inches(9, 108, forward=True)
    #group_stimfreq.draw(show=True)
    fig.savefig(str(path+ '\\id_Rheo.png'), dpi=300)

    mean_Rheo = (ggplot(data=datarheostat, mapping=aes(x='Group', y='Rheobase_mean', color='Group', ymin=0, ymax=1200))
               + geom_point(size=1)
               #                    + stat_summary(geom = "bar", fun_y = np.mean)
               #                    + stat_summary(geom = "errorbar", fun_data = 'mean_se')
               # + geom_boxplot()
                +geom_bar(aes(x='Group', y='Rheobase_mean'), stat='identity')
                +geom_errorbar(aes(x='Group', ymax='Rheobase_mean + Rheobase_sem', ymin='Rheobase_mean - Rheobase_sem'))
                +geom_point(aes(x='Group', y='Rheobase', color='CellID'), data=datarheo)
               #                    +geom_errorbar(aes(x='Group', y='Rheobase'), data=datarheoSEM)
               #                    + facet_grid('. ~ Genotype', space='free')
               # + theme_light()
#               + ylim(0, 1200)
               + theme(aspect_ratio=2 / 3)
               + theme(subplots_adjust={'right': .98, 'bottom': 0.25})
               + theme(strip_text_y=element_text(angle=0, ha='left'))
               + theme(axis_text_x=element_text(rotation=45, hjust=1)))
    fig = mean_Rheo.draw()
    # fig.set_size_inches(9, 108, forward=True)
    # group_stimfreq.draw(show=True)
    fig.savefig(str(path + '\\id_Rheo_Means.png'), dpi=300)

if bool_id_RheoInhib:
    mean_RheoInhibition = (ggplot(data=datarheostat, mapping=aes(x='Group', y='RheoInhibition_mean', color='Group', ymin=0, ymax=100))
                 + geom_point(size=1)
                 #                    + stat_summary(geom = "bar", fun_y = np.mean)
                 #                    + stat_summary(geom = "errorbar", fun_data = 'mean_se')
                 # + geom_boxplot()
                 + geom_bar(aes(x='Group', y='RheoInhibition_mean'), stat='identity')
                 + geom_errorbar(
                aes(x='Group', ymax='RheoInhibition_mean + RheoInhibition_sem', ymin='RheoInhibition_mean - RheoInhibition_sem'))
                 + geom_point(aes(x='Group', y='RheoInhibition', color='CellID'), data=datarheo)
                 #                    +geom_errorbar(aes(x='Group', y='Rheobase'), data=datarheoSEM)
                 #                    + facet_grid('. ~ Genotype', space='free')
                 # + theme_light()
                 #               + ylim(0, 1200)
                 + theme(aspect_ratio= 3/2)
                 + theme(subplots_adjust={'right': .98, 'bottom': 0.25})
                 + theme(strip_text_y=element_text(angle=0, ha='left'))
                 + theme(axis_text_x=element_text(rotation=45, hjust=1)))
    fig = mean_RheoInhibition.draw()
    # fig.set_size_inches(9, 108, forward=True)
    # group_stimfreq.draw(show=True)
    fig.savefig(str(path + '\\id_RheoInhibition_Means.png'), dpi=300)

if bool_id_RheoDiff:
    mean_RheoDiff = (ggplot(data=datarheostat,
                                  mapping=aes(x='Group', y='RheoDiff_mean', color='Group', ymin=0, ymax=1000))
                           + geom_point(size=1)
                           #                    + stat_summary(geom = "bar", fun_y = np.mean)
                           #                    + stat_summary(geom = "errorbar", fun_data = 'mean_se')
                           # + geom_boxplot()
                           + geom_bar(aes(x='Group', y='RheoDiff_mean'), stat='identity')
                           + geom_errorbar(
                aes(x='Group', ymax='RheoDiff_mean + RheoDiff_sem',
                    ymin='RheoDiff_mean - RheoDiff_sem'))
                           + geom_point(aes(x='Group', y='RheoDiff', color='CellID'), data=datarheo)
                           #                    +geom_errorbar(aes(x='Group', y='Rheobase'), data=datarheoSEM)
                           #                    + facet_grid('. ~ Genotype', space='free')
                           # + theme_light()
                           #               + ylim(0, 1200)
                           + theme(aspect_ratio=3/2)
                           + theme(subplots_adjust={'right': .98, 'bottom': 0.25})
                           + theme(strip_text_y=element_text(angle=0, ha='left'))
                           + theme(axis_text_x=element_text(rotation=45, hjust=1)))
    fig = mean_RheoDiff.draw()
    # fig.set_size_inches(9, 108, forward=True)
    # group_stimfreq.draw(show=True)
    fig.savefig(str(path + '\\id_RheoDiff_Means.png'), dpi=300)